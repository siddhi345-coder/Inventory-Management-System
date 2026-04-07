import os
import uuid
from io import BytesIO
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


TEMP_DIR = os.path.join(os.path.dirname(__file__), "..", "temp")
os.makedirs(TEMP_DIR, exist_ok=True)


def _apply_header_style(ws, header_fill_hex: str = "1E3A5F"):
    """Apply bold white text on dark background to the first row."""
    fill = PatternFill(start_color=header_fill_hex, end_color=header_fill_hex, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    border_side = Side(style="thin", color="CCCCCC")
    border = Border(
        left=border_side, right=border_side,
        top=border_side, bottom=border_side
    )
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _apply_row_styles(ws):
    """Alternate row colors and add borders for data rows."""
    light_fill = PatternFill(start_color="F0F4FA", end_color="F0F4FA", fill_type="solid")
    border_side = Side(style="thin", color="DDDDDD")
    border = Border(
        left=border_side, right=border_side,
        top=border_side, bottom=border_side
    )
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if i % 2 == 0:
                cell.fill = light_fill


def _auto_column_width(ws):
    """Auto-fit column widths based on content."""
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)


def generate_sales_excel(sales_data: list[dict]) -> str:
    """
    Generate a formatted Excel file from sales data.
    Returns the absolute file path of the saved .xlsx file.
    """
    if not sales_data:
        df = pd.DataFrame(columns=[
            "Sale ID", "Date", "Customer", "Product", "Brand",
            "Quantity", "Unit Price (₹)", "Line Total (₹)",
            "Sale Total (₹)", "Staff", "Status"
        ])
    else:
        df = pd.DataFrame(sales_data)
        df.rename(columns={
            "sale_id":       "Sale ID",
            "sale_date":     "Date",
            "customer_name": "Customer",
            "product_name":  "Product",
            "brand":         "Brand",
            "quantity":      "Quantity",
            "unit_price":    "Unit Price (₹)",
            "line_total":    "Line Total (₹)",
            "sale_total":    "Sale Total (₹)",
            "staff_name":    "Staff",
            "status":        "Status",
        }, inplace=True)

        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%d-%m-%Y %H:%M")

    file_name = f"sales_report_{uuid.uuid4().hex[:8]}.xlsx"
    file_path = os.path.join(TEMP_DIR, file_name)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sales Report")
        ws = writer.sheets["Sales Report"]

        # Style header row
        _apply_header_style(ws)
        _apply_row_styles(ws)
        _auto_column_width(ws)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add summary row at the bottom
        if not df.empty and "Line Total (₹)" in df.columns:
            summary_row = ws.max_row + 2
            ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
            total_col = df.columns.get_loc("Line Total (₹)") + 1
            total = df["Line Total (₹)"].sum()
            cell = ws.cell(row=summary_row, column=total_col, value=round(total, 2))
            cell.font = Font(bold=True, color="1E3A5F")

    return file_path
